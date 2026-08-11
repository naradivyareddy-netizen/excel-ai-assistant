# -*- coding: utf-8 -*-
"""
Excel AI Assistant Pro - WEB APP (Streamlit)
=============================================
Works in any browser: Windows, Mac, iPhone, iPad, Android.

Run locally:   streamlit run streamlit_app.py
Deploy free:   push this folder to GitHub -> share.streamlit.io
"""

import io
import os
import datetime

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px

try:
    import openpyxl
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False


# ----------------------------------------------------------
#  ERROR DEFINITIONS
# ----------------------------------------------------------
EXCEL_ERRORS = {
    "#VALUE!": "A value used in the formula is of the wrong data type "
               "(e.g. text where a number is expected).",
    "#DIV/0!": "The formula divides a number by zero or an empty cell.",
    "#N/A": "A lookup value could not be found (common with VLOOKUP / MATCH).",
    "#REF!": "A cell reference is invalid - usually because rows/columns "
             "were deleted.",
    "#NAME?": "Excel does not recognise text in the formula "
              "(misspelled function or missing quotes).",
    "#NUM!": "An invalid numeric value - number too large or wrong argument.",
    "#NULL!": "You used a space operator between ranges that do not intersect.",
}
ERROR_FIXES = {
    "#VALUE!": "Check all referenced cells contain numbers. Wrap text with "
               "VALUE() or use IFERROR(...,0).",
    "#DIV/0!": "Wrap the formula: =IFERROR(A1/B1,0) or test: =IF(B1=0,0,A1/B1).",
    "#N/A": "Use IFNA: =IFNA(VLOOKUP(...),\"Not found\"). Verify the lookup "
            "value exists.",
    "#REF!": "A reference was deleted. Rebuild the formula to a valid cell.",
    "#NAME?": "Fix the function spelling, add quotes around text, or define "
              "the named range.",
    "#NUM!": "Check arguments are valid (e.g. SQRT of a negative number).",
    "#NULL!": "Use a comma between ranges: SUM(A1:A5,B1:B5).",
}


# ----------------------------------------------------------
#  AI BRAIN (same offline logic as the desktop app)
# ----------------------------------------------------------
class AIBrain:
    @staticmethod
    def generate_formula(request):
        r = request.lower().strip()

        def block(f, e, s):
            return {"formula": f, "explanation": e, "sample": s}

        if any(k in r for k in ["tax", "gst", "vat"]):
            return block("=A2*B2   (Amount * Tax Rate)\nFor 18% GST: =A2*0.18\n"
                         "Total incl. tax: =A2*(1+0.18)",
                         "Multiply the taxable amount by the tax rate.",
                         "1000 @ 18% -> Tax 180, Total 1180")
        if "top" in r and any(c.isdigit() for c in r):
            return block("=LARGE($B$2:$B$1000, ROW()-1)\nOr: =SORT(A2:B1000,2,-1)",
                         "LARGE returns the k-th largest value; drag down. SORT "
                         "(Excel 365) sorts the whole table.",
                         "Top values listed highest to lowest.")
        if "lookup" in r or "vlookup" in r or "customer" in r:
            return block("=VLOOKUP(A2, Sheet2!$A$2:$D$1000, 2, FALSE)\n"
                         "Modern: =XLOOKUP(A2, Sheet2!A:A, Sheet2!B:B, \"Not found\")",
                         "VLOOKUP searches the first column and returns another "
                         "column. XLOOKUP is the safer modern version.",
                         "Look up customer ID -> return name.")
        if "attendance" in r:
            return block("Present: =COUNTIF(B2:B32,\"P\")\n"
                         "Absent:  =COUNTIF(B2:B32,\"A\")\n"
                         "Percent: =COUNTIF(B2:B32,\"P\")/COUNTA(B2:B32)",
                         "COUNTIF counts 'P' and 'A' marks, then computes %.",
                         "22 present of 26 -> 84.6%")
        if "average" in r or "mean" in r:
            return block("=AVERAGE(A2:A100)", "Mean of a range.", "Average value.")
        if "sum" in r or "total" in r:
            return block("=SUM(A2:A100)  /  =SUMIF(B2:B100,\"Sales\",A2:A100)",
                         "SUM totals a range; SUMIF totals matching rows.",
                         "Total of values / for one category.")
        if "count" in r:
            return block("=COUNTA(A2:A100)  /  =COUNTIF(A2:A100,\">100\")",
                         "COUNTA counts filled cells; COUNTIF counts matches.",
                         "Number of filled / matching cells.")
        if "percent" in r or "%" in r or "growth" in r:
            return block("=(B2-A2)/A2   (format as %)",
                         "Percent change = (New - Old) / Old.",
                         "200 to 250 -> 25%")
        if "if" in r or "condition" in r:
            return block("=IF(A2>=50,\"Pass\",\"Fail\")",
                         "IF returns one value if TRUE, another if FALSE.",
                         "72 -> Pass")
        if "date" in r or "day" in r or "month" in r:
            return block("Today: =TODAY()\nDays between: =B2-A2\n"
                         "Month name: =TEXT(A2,\"mmmm\")",
                         "Date functions for current date, gaps, formatting.",
                         "Jan 1 to Jan 31 -> 30 days")
        return block("=SUM(A2:A100)\n=AVERAGE(A2:A100)\n=IF(A2>0,\"Yes\",\"No\")",
                     "No exact match - here are useful starters. Try 'calculate "
                     "tax', 'top 10 sales', 'lookup customer'.",
                     "General purpose formulas.")

    @staticmethod
    def generate_vba(request):
        r = request.lower()
        if "attendance" in r:
            return ('Sub EmployeeAttendanceReport()\n'
                    '    Dim ws As Worksheet, lastRow As Long, i As Long\n'
                    '    Set ws = ActiveSheet\n'
                    '    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\n'
                    '    For i = 2 To lastRow\n'
                    '        ws.Cells(i, 3) = Application.CountIf(ws.Rows(i), "P")\n'
                    '    Next i\n'
                    '    MsgBox "Attendance report created!", vbInformation\n'
                    'End Sub')
        if "remove" in r and "duplicate" in r:
            return ('Sub RemoveDuplicatesAll()\n'
                    '    ActiveSheet.Range("A1").CurrentRegion.RemoveDuplicates _\n'
                    '        Columns:=1, Header:=xlYes\n'
                    '    MsgBox "Duplicates removed."\nEnd Sub')
        if "color" in r or "highlight" in r:
            return ('Sub HighlightErrors()\n'
                    '    Dim c As Range\n'
                    '    For Each c In ActiveSheet.UsedRange\n'
                    '        If IsError(c.Value) Then c.Interior.Color = '
                    'RGB(255,199,206)\n'
                    '    Next c\nEnd Sub')
        if "email" in r or "mail" in r:
            return ('Sub SendByEmail()\n'
                    '    Dim OutApp As Object, OutMail As Object\n'
                    '    Set OutApp = CreateObject("Outlook.Application")\n'
                    '    Set OutMail = OutApp.CreateItem(0)\n'
                    '    With OutMail\n'
                    '        .To = "friend@example.com"\n'
                    '        .Subject = "Report"\n'
                    '        .Attachments.Add ActiveWorkbook.FullName\n'
                    '        .Display\n    End With\nEnd Sub')
        return ('Sub CustomMacro()\n'
                '    Dim ws As Worksheet, lastRow As Long, i As Long\n'
                '    Set ws = ActiveSheet\n'
                '    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\n'
                '    For i = 2 To lastRow\n'
                "        ' your logic here\n"
                '        Debug.Print ws.Cells(i, 1).Value\n'
                '    Next i\nEnd Sub')

    @staticmethod
    def chat(message):
        m = message.lower()
        if "vlookup" in m and "xlookup" in m:
            return ("OLD: =VLOOKUP(A2,B:D,3,FALSE)\n"
                    "NEW: =XLOOKUP(A2,B:B,D:D,\"Not found\")\n\n"
                    "XLOOKUP can look left, gives a clean 'Not found', and "
                    "doesn't break when you insert columns.")
        if "vlookup" in m:
            return ("=VLOOKUP(lookup_value, table_range, column_number, FALSE)\n"
                    "Example: =VLOOKUP(A2, Sheet2!A:D, 2, FALSE)\n"
                    "Use FALSE for an exact match.")
        if "pivot" in m:
            return ("A Pivot Table summarises data without formulas.\n"
                    "1) Select data  2) Insert > PivotTable  3) Drag a category "
                    "to Rows and a number to Values. Or use the Pivot tab here.")
        if "power query" in m or "powerquery" in m:
            return ("Power Query: Data > Get Data > From File, then Remove "
                    "Columns / Change Type / Remove Duplicates, then Close & Load.")
        if "not working" in m or "error" in m or "wrong" in m:
            return ("Common causes: #DIV/0! (divide by empty), #N/A (lookup not "
                    "found), #VALUE! (text mixed with numbers), #NAME? (typo). "
                    "Upload your file and use the Formula Fixer tab.")
        if any(g in m for g in ["hello", "hi", "hey"]):
            return ("Hi! Ask me to make formulas, fix errors, explain pivots, "
                    "write VBA, or build a dashboard.")
        if "thank" in m:
            return "You're welcome! 😊"
        g = AIBrain.generate_formula(message)
        return "Here's a suggestion:\n\n%s\n\n%s" % (g["formula"], g["explanation"])


# ----------------------------------------------------------
#  WORKBOOK ANALYSIS (web version - works on uploaded bytes)
# ----------------------------------------------------------
def read_sheets(file_bytes, filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        return {"Sheet1": pd.read_csv(io.BytesIO(file_bytes))}
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    return {name: xls.parse(name) for name in xls.sheet_names}


def scan_errors(file_bytes, filename, sheets):
    errors = []
    seen = set()

    def add(sheet, cell, tok):
        key = (sheet, cell, tok)
        if key in seen:
            return
        seen.add(key)
        errors.append({"Sheet": sheet, "Location": cell, "Error": tok,
                       "Reason": EXCEL_ERRORS.get(tok, "Needs review"),
                       "Suggested Fix": ERROR_FIXES.get(tok, "Review manually")})

    is_xlsx = filename.lower().endswith(".xlsx")
    scanned = False
    if is_xlsx and HAVE_OPENPYXL:
        try:
            wbd = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for ws in wbd.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        tok = str(cell.value).strip()
                        if tok in EXCEL_ERRORS:
                            add(ws.title, cell.coordinate, tok)
            wbf = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
            for ws in wbf.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.startswith("="):
                            for tok in EXCEL_ERRORS:
                                if tok in v:
                                    add(ws.title, cell.coordinate, tok)
                            if cell.coordinate in v[1:]:
                                add(ws.title, cell.coordinate,
                                    "Circular Reference")
            scanned = True
        except Exception:
            scanned = False
    if not scanned:
        for name, df in sheets.items():
            for col in df.columns:
                for idx, val in df[col].items():
                    tok = str(val).strip()
                    if tok in EXCEL_ERRORS:
                        add(name, "%s row %s" % (col, idx + 2), tok)
    return errors


def auto_fix(file_bytes, filename):
    """Return (fixed_bytes, fixed_list, manual_list, out_ext)."""
    fixed, manual = [], []

    def wrap(formula, err):
        inner = formula[1:] if formula.startswith("=") else formula
        if err == "#N/A":
            return '=IFNA(%s,"Not found")' % inner
        return "=IFERROR(%s,0)" % inner

    is_xlsx = filename.lower().endswith(".xlsx")
    if is_xlsx and HAVE_OPENPYXL:
        wbd = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        wbf = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        for ws in wbf.worksheets:
            wsd = wbd[ws.title] if ws.title in wbd.sheetnames else None
            for row in ws.iter_rows():
                for cell in row:
                    coord = cell.coordinate
                    fval = cell.value
                    dval = wsd[coord].value if wsd is not None else None
                    cached = str(dval).strip() if dval is not None else ""
                    if isinstance(fval, str) and fval.startswith("="):
                        if coord in fval[1:]:
                            manual.append({"Sheet": ws.title, "Cell": coord,
                                           "Error": "Circular Reference"})
                            continue
                        err = cached if cached in EXCEL_ERRORS else None
                        if not err:
                            for tok in EXCEL_ERRORS:
                                if tok in fval:
                                    err = tok
                                    break
                        if err:
                            cell.value = wrap(fval, err)
                            fixed.append({"Sheet": ws.title, "Cell": coord,
                                          "Error": err,
                                          "Action": "wrapped with IFERROR"})
                    else:
                        sval = str(fval).strip() if fval is not None else ""
                        if sval in EXCEL_ERRORS:
                            cell.value = 0
                            fixed.append({"Sheet": ws.title, "Cell": coord,
                                          "Error": sval,
                                          "Action": "replaced with 0"})
        out = io.BytesIO()
        wbf.save(out)
        return out.getvalue(), fixed, manual, ".xlsx"

    # CSV / xls fallback via pandas
    sheets = read_sheets(file_bytes, filename)
    is_csv = filename.lower().endswith(".csv")
    for name, df in sheets.items():
        for col in df.columns:
            def _r(v, _n=name, _c=col):
                if str(v).strip() in EXCEL_ERRORS:
                    fixed.append({"Sheet": _n, "Cell": str(_c),
                                  "Error": str(v).strip(),
                                  "Action": "replaced with 0"})
                    return 0
                return v
            df[col] = df[col].map(_r)
    out = io.BytesIO()
    if is_csv:
        list(sheets.values())[0].to_csv(out, index=False)
        return out.getvalue(), fixed, manual, ".csv"
    with pd.ExcelWriter(out) as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name[:31], index=False)
    return out.getvalue(), fixed, manual, ".xlsx"


def clean_df(df, opts):
    report, original = [], len(df)
    if opts["duplicates"]:
        b = len(df); df = df.drop_duplicates()
        report.append("Removed %d duplicate rows." % (b - len(df)))
    if opts["blanks"]:
        b = len(df); df = df.dropna(how="all")
        report.append("Removed %d blank rows." % (b - len(df)))
    if opts["trim"]:
        for c in df.select_dtypes(include="object").columns:
            df[c] = df[c].astype(str).str.strip()
        report.append("Trimmed whitespace from text columns.")
    if opts["title"]:
        for c in df.select_dtypes(include="object").columns:
            df[c] = df[c].astype(str).str.title()
        report.append("Standardised text to Title Case.")
    if opts["dates"]:
        n = 0
        for c in df.columns:
            if "date" in str(c).lower():
                df[c] = pd.to_datetime(df[c], errors="coerce"); n += 1
        report.append("Fixed dates in %d column(s)." % n)
    if opts["columns"]:
        df.columns = [str(c).strip().title() for c in df.columns]
        report.append("Standardised column headers.")
    report.append("Missing values remaining: %d" % int(df.isna().sum().sum()))
    report.append("Rows: %d -> %d" % (original, len(df)))
    return df, report


def health_check(sheets, errors):
    total = max(sum(d.size for d in sheets.values()), 1)
    missing = sum(int(d.isna().sum().sum()) for d in sheets.values())
    dup = sum(int(d.duplicated().sum()) for d in sheets.values())
    err = len(errors)
    cols = sum(len(d.columns) for d in sheets.values())
    error_s = max(0, 100 - err * 8)
    quality_s = max(0, 100 - int(missing / total * 100) - int(dup / total * 50))
    simpl_s = max(0, 100 - cols * 2)
    perf_s = max(0, 100 - int(total / 5000))
    overall = int((error_s + quality_s + simpl_s + perf_s) / 4)
    tips = []
    if missing:
        tips.append("Fill or remove %d missing values." % missing)
    if dup:
        tips.append("Remove %d duplicate rows." % dup)
    if err:
        tips.append("Fix %d formula error(s)." % err)
    if not tips:
        tips.append("Your workbook looks healthy!")
    return {"overall": overall, "error": error_s, "quality": quality_s,
            "simplicity": simpl_s, "performance": perf_s,
            "missing": missing, "duplicates": dup, "errors": err, "tips": tips}


# ==========================================================
#  STREAMLIT UI
# ==========================================================
st.set_page_config(page_title="Excel AI Assistant Pro", page_icon="📊",
                   layout="wide")

st.markdown("""
<style>
.main .block-container { padding-top: 1.5rem; }
.big-title { font-size: 34px; font-weight: 800; color: #1b3a6b; }
.sub { color: #6b7c99; font-size: 15px; }
.stTabs [data-baseweb="tab"] { font-size: 15px; font-weight: 600; }
div.stButton > button { background:#2f80ed; color:white; border:none;
  border-radius:8px; padding:8px 18px; font-weight:700; }
div.stButton > button:hover { background:#1c6fe0; color:white; }
.kpi { background:linear-gradient(135deg,#2f80ed,#1b3a6b); color:white;
  padding:18px; border-radius:14px; text-align:center; }
.kpi h2 { margin:0; font-size:30px; } .kpi p { margin:0; opacity:.9; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="big-title">📊 Excel AI Assistant Pro</div>',
            unsafe_allow_html=True)
st.markdown('<div class="sub">Find, explain & auto-fix Excel errors — plus '
            'formulas, dashboards, VBA & reports. Works on any device.</div>',
            unsafe_allow_html=True)
st.write("")

# ----- File upload (shared across all tabs) -----
uploaded_list = st.file_uploader(
    "📁 Upload your Excel / CSV file (you can drop one or more — the first "
    "valid one is used)",
    type=["xlsx", "xls", "csv"], accept_multiple_files=True)

sheets, errors, fname, fbytes = {}, [], None, None
if uploaded_list:
    if len(uploaded_list) > 1:
        st.info("You added %d files — using the first one: **%s**"
                % (len(uploaded_list), uploaded_list[0].name))
    uploaded = uploaded_list[0]
    fname = uploaded.name
    fbytes = uploaded.getvalue()
    try:
        sheets = read_sheets(fbytes, fname)
        errors = scan_errors(fbytes, fname, sheets)
        st.success("✅ Loaded **%s** — %d worksheet(s), %d error(s) found."
                   % (fname, len(sheets), len(errors)))
    except Exception as e:
        st.error("Could not read **%s**: %s  —  make sure it is a valid "
                 ".xlsx, .xls or .csv file." % (fname, e))

# KPI row
if sheets:
    hc = health_check(sheets, errors)
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown('<div class="kpi"><h2>%d</h2><p>Worksheets</p></div>'
                % len(sheets), unsafe_allow_html=True)
    c2.markdown('<div class="kpi"><h2>%d</h2><p>Total Rows</p></div>'
                % sum(len(d) for d in sheets.values()), unsafe_allow_html=True)
    c3.markdown('<div class="kpi"><h2>%d</h2><p>Errors Found</p></div>'
                % len(errors), unsafe_allow_html=True)
    c4.markdown('<div class="kpi"><h2>%d</h2><p>Health Score</p></div>'
                % hc["overall"], unsafe_allow_html=True)
    st.write("")

tabs = st.tabs(["🛠 Formula Fixer", "🧹 Data Cleaner", "✨ Formula Generator",
                "📐 Pivot", "⚙️ VBA", "📈 Dashboard", "💬 AI Chat",
                "❤️ Health", "📄 Report", "👀 Preview"])

# ---- Formula Fixer ----
with tabs[0]:
    st.subheader("Formula Error Detection & Auto-Fix")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        st.markdown("**Step 1 — Review errors yourself:**")
        if errors:
            st.dataframe(pd.DataFrame(errors), use_container_width=True)
        else:
            st.success("No errors found — your workbook is clean! 🎉")
        st.markdown("**Step 2 — Auto-fix and download a corrected file:**")
        if st.button("⚡ Auto-Fix & Create Corrected File"):
            fixed_bytes, fixed, manual, ext = auto_fix(fbytes, fname)
            st.success("Fixed %d error(s)." % len(fixed))
            if fixed:
                st.dataframe(pd.DataFrame(fixed), use_container_width=True)
            if manual:
                st.warning("%d item(s) need manual review (e.g. circular "
                           "references):" % len(manual))
                st.dataframe(pd.DataFrame(manual), use_container_width=True)
            out_name = os.path.splitext(fname)[0] + "_FIXED" + ext
            st.download_button("⬇️ Download Fixed File", data=fixed_bytes,
                               file_name=out_name)

# ---- Data Cleaner ----
with tabs[1]:
    st.subheader("Data Cleaning")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        sheet = st.selectbox("Worksheet", list(sheets.keys()), key="clean_sheet")
        cols = st.columns(3)
        opts = {
            "duplicates": cols[0].checkbox("Remove duplicate rows", True),
            "blanks": cols[0].checkbox("Remove blank rows", True),
            "trim": cols[1].checkbox("Trim extra spaces", True),
            "title": cols[1].checkbox("Title Case text", False),
            "dates": cols[2].checkbox("Fix date formats", True),
            "columns": cols[2].checkbox("Standardise headers", True),
        }
        if st.button("🧹 Clean Now"):
            cleaned, report = clean_df(sheets[sheet].copy(), opts)
            for r in report:
                st.write("• " + r)
            st.dataframe(cleaned.head(100), use_container_width=True)
            out = io.BytesIO()
            cleaned.to_excel(out, index=False)
            st.download_button("⬇️ Download Cleaned File", data=out.getvalue(),
                               file_name="cleaned_data.xlsx")

# ---- Formula Generator ----
with tabs[2]:
    st.subheader("AI Formula Generator")
    req = st.text_input("Describe what you need",
                        placeholder="e.g. calculate tax, find top 10 sales")
    if st.button("✨ Generate Formula") and req:
        res = AIBrain.generate_formula(req)
        st.code(res["formula"], language="text")
        st.write("**Explanation:** " + res["explanation"])
        st.write("**Sample:** " + res["sample"])

# ---- Pivot ----
with tabs[3]:
    st.subheader("Pivot Table Assistant")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        sheet = st.selectbox("Worksheet", list(sheets.keys()), key="piv_sheet")
        df = sheets[sheet]
        c = st.columns(3)
        rk = c[0].selectbox("Group by", list(df.columns), key="piv_row")
        nums = list(df.select_dtypes(include=np.number).columns) or list(df.columns)
        vk = c[1].selectbox("Value", nums, key="piv_val")
        fn = c[2].selectbox("Function", ["sum", "mean", "count", "max", "min"],
                            key="piv_func")
        if st.button("📐 Build Pivot Table"):
            try:
                piv = pd.pivot_table(df, index=rk, values=vk,
                                     aggfunc=fn).reset_index()
                st.dataframe(piv, use_container_width=True)
            except Exception as e:
                st.error(str(e))

# ---- VBA ----
with tabs[4]:
    st.subheader("VBA / Macro Generator")
    vreq = st.text_input("Describe the macro",
                         placeholder="e.g. create employee attendance report")
    if st.button("⚙️ Generate VBA") and vreq:
        code = AIBrain.generate_vba(vreq)
        st.code(code, language="vbnet")
        st.caption("In Excel: press ALT+F11 → Insert > Module → paste → F5.")

# ---- Dashboard ----
with tabs[5]:
    st.subheader("Dashboard Generator")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        sheet = st.selectbox("Worksheet", list(sheets.keys()), key="dash_sheet")
        df = sheets[sheet]
        c = st.columns(2)
        cat = c[0].selectbox("Category", list(df.columns), key="dash_cat")
        nums = list(df.select_dtypes(include=np.number).columns) or list(df.columns)
        val = c[1].selectbox("Value", nums, key="dash_val")
        if st.button("📈 Generate Dashboard"):
            try:
                g = df.groupby(cat)[val].sum().sort_values(
                    ascending=False).head(10).reset_index()
                col1, col2 = st.columns(2)
                col1.plotly_chart(px.bar(g, x=cat, y=val,
                                  title="Top by %s" % val),
                                  use_container_width=True)
                col2.plotly_chart(px.pie(g.head(5), names=cat, values=val,
                                  title="Share (Top 5)"),
                                  use_container_width=True)
                st.plotly_chart(px.line(g, x=cat, y=val, markers=True,
                                title="Trend"), use_container_width=True)
                st.metric("Total %s" % val, "{:,.0f}".format(g[val].sum()))
            except Exception as e:
                st.error(str(e))

# ---- AI Chat ----
with tabs[6]:
    st.subheader("Ask the Excel AI")
    if "chat" not in st.session_state:
        st.session_state.chat = [("AI", "Hi! Ask me about formulas, errors, "
                                  "VLOOKUP, pivots, VBA and more.")]
    q = st.chat_input("Type your question...")
    if q:
        st.session_state.chat.append(("You", q))
        st.session_state.chat.append(("AI", AIBrain.chat(q)))
    for who, msg in st.session_state.chat:
        with st.chat_message("user" if who == "You" else "assistant"):
            st.write(msg)

# ---- Health ----
with tabs[7]:
    st.subheader("Workbook Health Check")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        hc = health_check(sheets, errors)
        for key, label in [("overall", "Overall"), ("error", "Error Score"),
                           ("quality", "Data Quality"),
                           ("simplicity", "Formula Simplicity"),
                           ("performance", "Performance")]:
            st.write("**%s** — %d/100" % (label, hc[key]))
            st.progress(hc[key] / 100)
        st.write("**Suggestions:**")
        for t in hc["tips"]:
            st.write("• " + t)

# ---- Report ----
with tabs[8]:
    st.subheader("Audit Report")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        hc = health_check(sheets, errors)
        lines = ["EXCEL AI ASSISTANT PRO - AUDIT REPORT",
                 "Generated: " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                 "File: " + (fname or "-"), "",
                 "STRUCTURE:"]
        for n, d in sheets.items():
            lines.append("  %s: %d rows x %d cols" % (n, len(d), len(d.columns)))
        lines += ["", "HEALTH:",
                  "  Overall %d | Error %d | Quality %d | Simplicity %d | "
                  "Performance %d" % (hc["overall"], hc["error"], hc["quality"],
                                      hc["simplicity"], hc["performance"]),
                  "", "ERRORS (%d):" % len(errors)]
        for e in errors:
            lines.append("  [%s] %s @ %s" % (e["Error"], e["Sheet"],
                                             e["Location"]))
        lines += ["", "RECOMMENDATIONS:"] + ["  • " + t for t in hc["tips"]]
        report = "\n".join(lines)
        st.text(report)
        st.download_button("⬇️ Download Report (.txt)", data=report,
                           file_name="audit_report.txt")

# ---- Preview ----
with tabs[9]:
    st.subheader("Data Preview")
    if not sheets:
        st.info("⬆️ Upload a file above to begin.")
    else:
        sheet = st.selectbox("Worksheet", list(sheets.keys()), key="prev_sheet")
        st.dataframe(sheets[sheet].head(300), use_container_width=True)

st.write("---")
st.caption("Excel AI Assistant Pro • Web Edition • Works on iPhone, Mac & Windows")
