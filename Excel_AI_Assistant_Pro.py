# -*- coding: utf-8 -*-
"""
============================================================
        EXCEL AI ASSISTANT PRO  -  Single File Edition
============================================================
A professional, AI-powered Excel troubleshooting & productivity
desktop assistant.

HOW TO RUN
----------
1. Just open this file in IDLE (or any Python editor) and press F5 / Run.
2. On the FIRST run it will automatically install every library it needs.
   (Please be patient the first time - it may take a few minutes.)
3. After install it launches the app window automatically.

You do NOT need to install anything yourself. One file. One click.

Author: Senior Python Developer
============================================================
"""

# ----------------------------------------------------------
#  STEP 1 :  AUTO-INSTALL EVERYTHING  (no "No module" errors)
# ----------------------------------------------------------
import sys
import subprocess
import importlib


def _install_requirements():
    """Install all required packages quietly before the app starts."""
    # When packaged as a standalone .exe, all libraries are already bundled.
    if getattr(sys, "frozen", False):
        return
    # package on pip  ->  module name to import-test
    required = {
        "PyQt5": "PyQt5",
        "pandas": "pandas",
        "numpy": "numpy",
        "openpyxl": "openpyxl",
        "xlrd": "xlrd",
        "matplotlib": "matplotlib",
        "pyttsx3": "pyttsx3",
    }

    # Upgrade pip silently (best-effort, never blocks the app)
    try:
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--upgrade", "pip"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
    except Exception:
        pass

    for pip_name, module_name in required.items():
        try:
            importlib.import_module(module_name)
        except Exception:
            print("Installing %s ... (first run only)" % pip_name)
            try:
                subprocess.run(
                    [sys.executable, "-m", "pip", "install", pip_name],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
            except Exception as exc:
                print("Could not auto-install %s: %s" % (pip_name, exc))


_install_requirements()

# ----------------------------------------------------------
#  STEP 2 :  IMPORTS (safe, with graceful fallbacks)
# ----------------------------------------------------------
import os
import re
import datetime
import traceback

import numpy as np
import pandas as pd

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

# Text-to-speech (voice output) - optional, never crashes the app
try:
    import pyttsx3
    HAVE_TTS = True
except Exception:
    HAVE_TTS = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QStackedWidget, QFileDialog, QTableWidget, QTableWidgetItem,
    QTextEdit, QLineEdit, QComboBox, QFrame, QScrollArea, QPlainTextEdit,
    QListWidget, QListWidgetItem, QMessageBox, QSizePolicy, QSpacerItem,
    QHeaderView, QProgressBar, QGridLayout, QTabWidget, QCheckBox
)
from PyQt5.QtGui import QFont, QIcon, QColor, QPalette, QClipboard
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal, QTimer

# Matplotlib embedded in Qt
import matplotlib
matplotlib.use("Qt5Agg")
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


APP_NAME = "Excel AI Assistant Pro"
APP_VERSION = "1.0"

# Excel error tokens we look for
EXCEL_ERRORS = {
    "#VALUE!": "A value used in the formula is of the wrong data type "
               "(e.g. text where a number is expected).",
    "#DIV/0!": "The formula divides a number by zero or an empty cell.",
    "#N/A": "A lookup value could not be found (common with VLOOKUP / MATCH).",
    "#REF!": "A cell reference is invalid - usually because rows/columns "
             "were deleted.",
    "#NAME?": "Excel does not recognise text in the formula "
              "(misspelled function or missing quotes).",
    "#NUM!": "An invalid numeric value - number too large or wrong argument.",
    "#NULL!": "You used a space (intersection) operator between ranges that "
              "do not intersect.",
}

ERROR_FIXES = {
    "#VALUE!": "Check that all referenced cells contain numbers. Wrap text "
               "values with VALUE() or use IFERROR(...,0).",
    "#DIV/0!": "Wrap the formula: =IFERROR(A1/B1,0)  or test the divisor: "
               "=IF(B1=0,0,A1/B1).",
    "#N/A": "Use IFNA: =IFNA(VLOOKUP(...),\"Not found\"). Verify the lookup "
            "value exists and ranges match.",
    "#REF!": "A reference was deleted. Rebuild the formula and point it to a "
             "valid cell/range.",
    "#NAME?": "Fix the function spelling, add missing quotes around text, or "
              "define the named range.",
    "#NUM!": "Check arguments are within valid range (e.g. SQRT of a negative "
             "number, or too-large results).",
    "#NULL!": "Use a comma between ranges instead of a space, e.g. SUM(A1:A5,"
              "B1:B5).",
}


# ==========================================================
#  AI / LOGIC ENGINE  (offline rule-based "brain")
# ==========================================================
class AIBrain:
    """A lightweight offline assistant that answers Excel questions,
    builds formulas, VBA macros, and explanations - no API key needed."""

    # ---------- FORMULA GENERATOR ----------
    @staticmethod
    def generate_formula(request):
        r = request.lower().strip()

        def block(formula, explanation, sample):
            return {"formula": formula, "explanation": explanation,
                    "sample": sample}

        if any(k in r for k in ["tax", "gst", "vat"]):
            return block(
                "=A2*B2     (Amount * Tax Rate)\n"
                "Example for 18% GST:  =A2*0.18\n"
                "Total incl. tax:      =A2*(1+0.18)",
                "Multiply the taxable amount by the tax rate. Put the amount "
                "in column A and the rate in column B.",
                "Amount 1000 @ 18%  ->  Tax = 180,  Total = 1180")

        if "top" in r and any(c.isdigit() for c in r):
            n = re.findall(r"\d+", r)
            n = n[0] if n else "10"
            return block(
                "=LARGE($B$2:$B$1000, ROW()-1)\n"
                "Or use:  =SORT(A2:B1000,2,-1)   (Excel 365)",
                "LARGE returns the k-th largest value. Drag down %s rows to "
                "get the top %s. SORT (Excel 365) sorts the whole table." % (n, n),
                "Top sales values listed from highest to lowest.")

        if "lookup" in r or "vlookup" in r or "customer" in r:
            return block(
                "=VLOOKUP(A2, Sheet2!$A$2:$D$1000, 2, FALSE)\n"
                "Modern:  =XLOOKUP(A2, Sheet2!A:A, Sheet2!B:B, \"Not found\")",
                "VLOOKUP searches the first column of a range and returns a "
                "value from another column. XLOOKUP is the newer, safer "
                "version and works in any direction.",
                "Look up customer ID -> return customer name.")

        if "attendance" in r:
            return block(
                "Present count:  =COUNTIF(B2:B32,\"P\")\n"
                "Absent count:   =COUNTIF(B2:B32,\"A\")\n"
                "Attendance %:   =COUNTIF(B2:B32,\"P\")/COUNTA(B2:B32)",
                "Use COUNTIF to count 'P' (present) and 'A' (absent) marks "
                "for each person, then compute the percentage.",
                "22 present out of 26 days  ->  84.6%")

        if "average" in r or "mean" in r:
            return block("=AVERAGE(A2:A100)",
                         "AVERAGE returns the arithmetic mean of a range.",
                         "Average of the selected numbers.")

        if "sum" in r or "total" in r:
            return block("=SUM(A2:A100)   or   =SUMIF(B2:B100,\"Sales\",A2:A100)",
                         "SUM totals a range. SUMIF totals only rows that "
                         "match a condition.",
                         "Total of all values / total for one category.")

        if "count" in r:
            return block("=COUNTA(A2:A100)   /   =COUNTIF(A2:A100,\">100\")",
                         "COUNTA counts non-empty cells. COUNTIF counts cells "
                         "meeting a condition.",
                         "Number of filled cells or matching cells.")

        if "percent" in r or "%" in r or "growth" in r:
            return block("=(B2-A2)/A2     (format the cell as %)",
                         "Percentage change = (New - Old) / Old.",
                         "From 200 to 250  ->  25% growth")

        if "if" in r or "condition" in r:
            return block("=IF(A2>=50,\"Pass\",\"Fail\")",
                         "IF returns one value when the test is TRUE and "
                         "another when FALSE.",
                         "Score 72  ->  \"Pass\"")

        if "date" in r or "day" in r or "month" in r:
            return block(
                "Today:        =TODAY()\n"
                "Days between: =B2-A2\n"
                "Month name:   =TEXT(A2,\"mmmm\")",
                "Date functions for current date, differences and formatting.",
                "01-Jan to 31-Jan  ->  30 days")

        # default
        return block(
            "=SUM(A2:A100)\n=AVERAGE(A2:A100)\n=IF(A2>0,\"Yes\",\"No\")",
            "I couldn't match an exact template, so here are useful starters. "
            "Try phrasing like 'calculate tax', 'top 10 sales', "
            "'lookup customer', or 'attendance percentage'.",
            "General purpose formulas.")

    # ---------- VBA GENERATOR ----------
    @staticmethod
    def generate_vba(request):
        r = request.lower()

        if "attendance" in r:
            return ('Sub EmployeeAttendanceReport()\n'
                    '    Dim ws As Worksheet, rep As Worksheet\n'
                    '    Dim lastRow As Long, i As Long, present As Long\n'
                    '    Set ws = ActiveSheet\n'
                    '    On Error Resume Next\n'
                    '    Set rep = Sheets("Attendance Report")\n'
                    '    On Error GoTo 0\n'
                    '    If rep Is Nothing Then Set rep = Sheets.Add: '
                    'rep.Name = "Attendance Report"\n'
                    '    rep.Cells.Clear\n'
                    '    rep.Range("A1:C1").Value = Array("Employee", '
                    '"Present", "Percent")\n'
                    '    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\n'
                    '    For i = 2 To lastRow\n'
                    '        present = Application.CountIf(ws.Rows(i), "P")\n'
                    '        rep.Cells(i, 1) = ws.Cells(i, 1).Value\n'
                    '        rep.Cells(i, 2) = present\n'
                    '        rep.Cells(i, 3) = Format(present / '
                    '(lastRow - 1), "0.0%")\n'
                    '    Next i\n'
                    '    MsgBox "Attendance report created!", vbInformation\n'
                    'End Sub')

        if "remove" in r and "duplicate" in r:
            return ('Sub RemoveDuplicates_AllColumns()\n'
                    '    Dim ws As Worksheet\n'
                    '    Set ws = ActiveSheet\n'
                    '    ws.Range("A1").CurrentRegion.RemoveDuplicates _\n'
                    '        Columns:=Application.Evaluate("ROW(1:" & _\n'
                    '        ws.Range("A1").CurrentRegion.Columns.Count & ")"), '
                    '_\n        Header:=xlYes\n'
                    '    MsgBox "Duplicates removed.", vbInformation\n'
                    'End Sub')

        if "color" in r or "highlight" in r:
            return ('Sub HighlightErrors()\n'
                    '    Dim c As Range\n'
                    '    For Each c In ActiveSheet.UsedRange\n'
                    '        If IsError(c.Value) Then\n'
                    '            c.Interior.Color = RGB(255, 199, 206)\n'
                    '        End If\n'
                    '    Next c\n'
                    '    MsgBox "All error cells highlighted in red.", '
                    'vbInformation\n'
                    'End Sub')

        if "email" in r or "mail" in r:
            return ('Sub SendWorkbookByEmail()\n'
                    '    Dim OutApp As Object, OutMail As Object\n'
                    '    Set OutApp = CreateObject("Outlook.Application")\n'
                    '    Set OutMail = OutApp.CreateItem(0)\n'
                    '    With OutMail\n'
                    '        .To = "recipient@example.com"\n'
                    '        .Subject = "Report - " & Format(Date, "dd-mmm")\n'
                    '        .Body = "Please find the report attached."\n'
                    '        .Attachments.Add ActiveWorkbook.FullName\n'
                    '        .Display   ' + chr(39) +'use .Send to send '
                    'automatically\n'
                    '    End With\n'
                    'End Sub')

        # default generic loop template
        return ('Sub CustomMacro()\n'
                '    Dim ws As Worksheet, lastRow As Long, i As Long\n'
                '    Set ws = ActiveSheet\n'
                '    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\n'
                '    For i = 2 To lastRow\n'
                "        ' TODO: your logic here for each row\n"
                '        Debug.Print ws.Cells(i, 1).Value\n'
                '    Next i\n'
                '    MsgBox "Macro finished!", vbInformation\n'
                'End Sub')

    # ---------- CHAT ASSISTANT ----------
    @staticmethod
    def chat(message):
        m = message.lower()

        if "vlookup" in m and "xlookup" in m:
            return ("Convert VLOOKUP to XLOOKUP:\n\n"
                    "OLD:  =VLOOKUP(A2, B:D, 3, FALSE)\n"
                    "NEW:  =XLOOKUP(A2, B:B, D:D, \"Not found\")\n\n"
                    "XLOOKUP is better because it can look left, returns a "
                    "clean 'Not found' message, and doesn't break when you "
                    "insert columns.")

        if "vlookup" in m:
            return ("VLOOKUP syntax:\n"
                    "=VLOOKUP(lookup_value, table_range, column_number, FALSE)\n\n"
                    "Example: =VLOOKUP(A2, Sheet2!A:D, 2, FALSE)\n"
                    "Always use FALSE for an exact match. Keep the lookup "
                    "column as the FIRST column of your range.")

        if "pivot" in m:
            return ("A Pivot Table summarises large data without formulas.\n"
                    "1) Select your data.\n"
                    "2) Insert > PivotTable.\n"
                    "3) Drag a category to Rows and a number to Values.\n"
                    "Excel instantly totals/averages by group. Great for "
                    "sales-by-region, expense-by-month, etc.\n\n"
                    "Tip: use this app's 'Pivot Assistant' tab to auto-build "
                    "one from your uploaded file.")

        if "power query" in m or "powerquery" in m:
            return ("Power Query steps to clean & combine data:\n"
                    "1) Data > Get Data > From File.\n"
                    "2) In the editor: Remove Columns, Change Type, "
                    "Remove Duplicates, Replace Values.\n"
                    "3) Home > Close & Load.\n"
                    "Every step is recorded and re-runs automatically when "
                    "data refreshes.")

        if "not working" in m or "error" in m or "wrong" in m:
            return ("Common reasons a formula fails:\n"
                    "- #DIV/0!  -> dividing by an empty/zero cell.\n"
                    "- #N/A     -> lookup value not found.\n"
                    "- #VALUE!  -> text mixed with numbers.\n"
                    "- #NAME?   -> misspelled function name.\n\n"
                    "Upload your file in the 'Upload Excel' tab and run "
                    "'Formula Fixer' - I'll point to the exact cells.")

        if "hello" in m or "hi" in m or "hey" in m:
            return ("Hello! I'm your Excel AI Assistant. Ask me to create "
                    "formulas, fix errors, explain pivot tables, write VBA, "
                    "or build a dashboard. How can I help?")

        if "thank" in m:
            return "You're welcome! Happy spreadsheeting. 😊"

        # try the formula generator as a fallback
        gen = AIBrain.generate_formula(message)
        return ("Here's what I can suggest:\n\n%s\n\n%s"
                % (gen["formula"], gen["explanation"]))


# ==========================================================
#  WORKBOOK ANALYSER
# ==========================================================
class Workbook:
    """Holds the loaded file and provides analysis helpers."""

    def __init__(self):
        self.path = None
        self.sheets = {}          # name -> DataFrame
        self.errors = []          # list of dicts

    def load(self, path):
        self.path = path
        self.sheets = {}
        self.errors = []
        ext = os.path.splitext(path)[1].lower()

        if ext == ".csv":
            df = pd.read_csv(path)
            self.sheets["Sheet1"] = df
        else:
            xls = pd.ExcelFile(path)
            for name in xls.sheet_names:
                self.sheets[name] = xls.parse(name)

        self._scan_errors()
        return self.sheets

    def _scan_errors(self):
        """Scan every cell value for Excel error tokens.

        NOTE: pandas converts Excel error cells (#DIV/0!, #N/A, #VALUE! ...)
        to NaN on read, so for .xlsx/.xls we scan the raw cells with openpyxl
        which preserves the error text. The dataframe scan is the fallback for
        CSV files and any error tokens that survive as plain text.
        """
        self.errors = []
        seen = set()

        def add(sheet, cell, token, reason, fix):
            key = (sheet, cell, token)
            if key in seen:
                return
            seen.add(key)
            self.errors.append({"sheet": sheet, "cell": cell, "error": token,
                                "reason": reason, "fix": fix})

        is_csv = bool(self.path and self.path.lower().endswith(".csv"))

        # ---- Raw cell scan via openpyxl (reliable for .xlsx) ----
        scanned_with_openpyxl = False
        if HAVE_OPENPYXL and self.path and not is_csv:
            try:
                wb = openpyxl.load_workbook(self.path, data_only=True)
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            token = str(cell.value).strip()
                            if token in EXCEL_ERRORS:
                                add(ws.title, cell.coordinate, token,
                                    EXCEL_ERRORS[token], ERROR_FIXES[token])
                # second pass on formulas for broken refs / circular loops
                wbf = openpyxl.load_workbook(self.path, data_only=False)
                for ws in wbf.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            v = cell.value
                            if isinstance(v, str) and v.startswith("="):
                                for tok in EXCEL_ERRORS:
                                    if tok in v:
                                        add(ws.title, cell.coordinate, tok,
                                            EXCEL_ERRORS[tok], ERROR_FIXES[tok])
                                body = v[1:]
                                if cell.coordinate in body:
                                    add(ws.title, cell.coordinate,
                                        "Circular Reference",
                                        "The formula refers to its own cell, "
                                        "creating a loop.",
                                        "Remove the self-reference or enable "
                                        "iterative calculation (File > Options "
                                        "> Formulas).")
                scanned_with_openpyxl = True
            except Exception:
                scanned_with_openpyxl = False

        # ---- DataFrame scan (CSV, .xls, or openpyxl fallback) ----
        if is_csv or not scanned_with_openpyxl:
            for sheet, df in self.sheets.items():
                for col in df.columns:
                    for idx, val in df[col].items():
                        token = str(val).strip()
                        if token in EXCEL_ERRORS:
                            add(sheet, "%s row %s" % (col, idx + 2), token,
                                EXCEL_ERRORS[token], ERROR_FIXES[token])
        return self.errors

    # ---------- AUTO-FIX & SAVE ----------
    def auto_fix(self, save_path):
        """Automatically repair errors and write a corrected file.

        - Formula cells that produce an error are wrapped so the error is
          neutralised:  =A/B  ->  =IFERROR(A/B,0)  ;  VLOOKUP -> IFNA(...).
        - Literal error values (typed text or cached errors) become 0.
        - Circular references are reported for manual review (cannot be
          auto-wrapped safely).

        Returns (fixed_list, manual_list).
        """
        fixed, manual = [], []

        def wrap(formula, err):
            inner = formula[1:] if formula.startswith("=") else formula
            if err == "#N/A":
                return '=IFNA(%s,"Not found")' % inner
            return "=IFERROR(%s,0)" % inner

        is_csv = bool(self.path and self.path.lower().endswith(".csv"))
        is_xlsx = bool(self.path and self.path.lower().endswith(".xlsx"))

        # ---- Best path: real .xlsx via openpyxl (preserves & wraps formulas) ----
        if is_xlsx and HAVE_OPENPYXL:
            wb_data = openpyxl.load_workbook(self.path, data_only=True)
            wb_f = openpyxl.load_workbook(self.path, data_only=False)
            for ws in wb_f.worksheets:
                ws_d = wb_data[ws.title] if ws.title in wb_data.sheetnames \
                    else None
                for row in ws.iter_rows():
                    for cell in row:
                        coord = cell.coordinate
                        fval = cell.value
                        dval = ws_d[coord].value if ws_d is not None else None
                        cached = str(dval).strip() if dval is not None else ""

                        if isinstance(fval, str) and fval.startswith("="):
                            # circular reference -> manual
                            if coord in fval[1:]:
                                manual.append({"sheet": ws.title, "cell": coord,
                                               "error": "Circular Reference"})
                                continue
                            err = None
                            if cached in EXCEL_ERRORS:
                                err = cached
                            else:
                                for tok in EXCEL_ERRORS:
                                    if tok in fval:
                                        err = tok
                                        break
                            if err:
                                cell.value = wrap(fval, err)
                                fixed.append({"sheet": ws.title, "cell": coord,
                                              "error": err,
                                              "action": "wrapped with IFERROR"})
                        else:
                            sval = str(fval).strip() if fval is not None else ""
                            if sval in EXCEL_ERRORS:
                                cell.value = 0
                                fixed.append({"sheet": ws.title, "cell": coord,
                                              "error": sval,
                                              "action": "replaced with 0"})
            if not save_path.lower().endswith(".xlsx"):
                save_path = os.path.splitext(save_path)[0] + ".xlsx"
            wb_f.save(save_path)
            return fixed, manual

        # ---- Fallback: CSV / .xls via pandas (replace error tokens with 0) ----
        new_sheets = {}
        for name, df in self.sheets.items():
            df2 = df.copy()
            for col in df2.columns:
                def _repl(v):
                    if str(v).strip() in EXCEL_ERRORS:
                        fixed.append({"sheet": name, "cell": str(col),
                                      "error": str(v).strip(),
                                      "action": "replaced with 0"})
                        return 0
                    return v
                df2[col] = df2[col].map(_repl)
            new_sheets[name] = df2

        if is_csv or save_path.lower().endswith(".csv"):
            list(new_sheets.values())[0].to_csv(save_path, index=False)
        else:
            with pd.ExcelWriter(save_path) as writer:
                for name, df2 in new_sheets.items():
                    df2.to_excel(writer, sheet_name=name[:31], index=False)
        return fixed, manual

    # ---------- DATA CLEANING ----------
    def clean(self, df, options):
        report = []
        original = len(df)

        if options.get("duplicates"):
            before = len(df)
            df = df.drop_duplicates()
            report.append("Removed %d duplicate rows." % (before - len(df)))

        if options.get("blanks"):
            before = len(df)
            df = df.dropna(how="all")
            report.append("Removed %d fully blank rows." % (before - len(df)))

        if options.get("trim"):
            obj_cols = df.select_dtypes(include="object").columns
            for c in obj_cols:
                df[c] = df[c].astype(str).str.strip()
            report.append("Trimmed whitespace in %d text columns."
                          % len(obj_cols))

        if options.get("title"):
            obj_cols = df.select_dtypes(include="object").columns
            for c in obj_cols:
                df[c] = df[c].astype(str).str.title()
            report.append("Standardised text to Title Case.")

        if options.get("dates"):
            count = 0
            for c in df.columns:
                if "date" in str(c).lower():
                    try:
                        df[c] = pd.to_datetime(df[c], errors="coerce")
                        count += 1
                    except Exception:
                        pass
            report.append("Fixed date format in %d column(s)." % count)

        if options.get("columns"):
            df.columns = [str(c).strip().title().replace("  ", " ")
                          for c in df.columns]
            report.append("Standardised column headers.")

        missing = int(df.isna().sum().sum())
        report.append("Detected %d missing value(s) remaining." % missing)
        report.append("Rows: %d  ->  %d" % (original, len(df)))
        return df, report

    # ---------- HEALTH CHECK ----------
    def health_check(self):
        if not self.sheets:
            return None
        total_cells = sum(df.size for df in self.sheets.values())
        total_cells = max(total_cells, 1)
        missing = sum(int(df.isna().sum().sum()) for df in self.sheets.values())
        dup = sum(int(df.duplicated().sum()) for df in self.sheets.values())
        err = len(self.errors)

        error_score = max(0, 100 - err * 8)
        quality_score = max(0, 100 - int((missing / total_cells) * 100)
                            - int((dup / max(total_cells, 1)) * 50))
        # formula complexity (rough): more sheets/cols => more complex
        cols = sum(len(df.columns) for df in self.sheets.values())
        complexity_score = max(0, 100 - cols * 2)
        perf_score = max(0, 100 - int(total_cells / 5000))

        overall = int((error_score + quality_score +
                       complexity_score + perf_score) / 4)

        tips = []
        if missing:
            tips.append("Fill or remove %d missing values." % missing)
        if dup:
            tips.append("Remove %d duplicate rows." % dup)
        if err:
            tips.append("Fix %d formula error(s) in the Formula Fixer tab." % err)
        if total_cells > 100000:
            tips.append("Large workbook - consider Power Query or splitting "
                        "sheets for speed.")
        if not tips:
            tips.append("Your workbook looks healthy. Great job!")

        return {
            "overall": overall,
            "error": error_score,
            "quality": quality_score,
            "complexity": complexity_score,
            "performance": perf_score,
            "tips": tips,
            "stats": {"cells": total_cells, "missing": missing,
                      "duplicates": dup, "errors": err},
        }


# ==========================================================
#  THEMES / STYLES
# ==========================================================
LIGHT_QSS = """
QMainWindow, QWidget { background: #f4f7fb; color: #1f2a44; }
#Sidebar { background: #1b3a6b; }
#SidebarTitle { color: #ffffff; font-size: 17px; font-weight: bold;
                padding: 18px 14px; }
#NavBtn { color: #d6e4ff; text-align: left; padding: 11px 18px;
          border: none; font-size: 14px; border-radius: 8px; }
#NavBtn:hover { background: #28508f; color: #ffffff; }
#NavBtn:checked { background: #2f80ed; color: #ffffff; font-weight: bold; }
#Header { background: #ffffff; border-bottom: 2px solid #e3e9f2; }
#HeaderTitle { font-size: 20px; font-weight: bold; color: #1b3a6b; }
#Card { background: #ffffff; border: 1px solid #e3e9f2; border-radius: 14px; }
#KpiCard { border-radius: 14px; }
QPushButton#Primary { background: #2f80ed; color: white; border: none;
    padding: 10px 18px; border-radius: 9px; font-weight: bold; font-size: 14px; }
QPushButton#Primary:hover { background: #1c6fe0; }
QPushButton#Success { background: #27ae60; color: white; border: none;
    padding: 10px 18px; border-radius: 9px; font-weight: bold; font-size: 14px; }
QPushButton#Success:hover { background: #1f9b53; }
QPushButton#Ghost { background: #eaf1fb; color: #1b3a6b; border: none;
    padding: 9px 16px; border-radius: 9px; font-weight: bold; }
QPushButton#Ghost:hover { background: #d8e6fb; }
QTextEdit, QPlainTextEdit, QLineEdit, QComboBox {
    background: #ffffff; border: 1px solid #cdd8e8; border-radius: 8px;
    padding: 8px; color: #1f2a44; font-size: 13px; }
QTableWidget { background: #ffffff; border: 1px solid #e3e9f2;
    border-radius: 8px; gridline-color: #e8eef7; }
QHeaderView::section { background: #1b3a6b; color: white; padding: 6px;
    border: none; font-weight: bold; }
QProgressBar { border: none; border-radius: 8px; background: #e3e9f2;
    height: 16px; text-align: center; color:#1f2a44; }
QProgressBar::chunk { border-radius: 8px; background: #2f80ed; }
QScrollArea { border: none; }
"""

DARK_QSS = """
QMainWindow, QWidget { background: #0f1626; color: #e6ecf5; }
#Sidebar { background: #0a0f1c; }
#SidebarTitle { color: #ffffff; font-size: 17px; font-weight: bold;
                padding: 18px 14px; }
#NavBtn { color: #9fb3d1; text-align: left; padding: 11px 18px;
          border: none; font-size: 14px; border-radius: 8px; }
#NavBtn:hover { background: #182238; color: #ffffff; }
#NavBtn:checked { background: #2f80ed; color: #ffffff; font-weight: bold; }
#Header { background: #131c30; border-bottom: 2px solid #1d2942; }
#HeaderTitle { font-size: 20px; font-weight: bold; color: #6fa8ff; }
#Card { background: #131c30; border: 1px solid #1d2942; border-radius: 14px; }
QPushButton#Primary { background: #2f80ed; color: white; border: none;
    padding: 10px 18px; border-radius: 9px; font-weight: bold; font-size: 14px; }
QPushButton#Primary:hover { background: #1c6fe0; }
QPushButton#Success { background: #27ae60; color: white; border: none;
    padding: 10px 18px; border-radius: 9px; font-weight: bold; font-size: 14px; }
QPushButton#Success:hover { background: #1f9b53; }
QPushButton#Ghost { background: #1d2942; color: #cfe0ff; border: none;
    padding: 9px 16px; border-radius: 9px; font-weight: bold; }
QPushButton#Ghost:hover { background: #28395c; }
QTextEdit, QPlainTextEdit, QLineEdit, QComboBox {
    background: #0d1422; border: 1px solid #25324d; border-radius: 8px;
    padding: 8px; color: #e6ecf5; font-size: 13px; }
QTableWidget { background: #0d1422; border: 1px solid #1d2942;
    border-radius: 8px; gridline-color: #1d2942; color:#e6ecf5; }
QHeaderView::section { background: #1b2740; color: #cfe0ff; padding: 6px;
    border: none; font-weight: bold; }
QProgressBar { border: none; border-radius: 8px; background: #1d2942;
    height: 16px; text-align: center; color:#e6ecf5; }
QProgressBar::chunk { border-radius: 8px; background: #2f80ed; }
QScrollArea { border: none; }
"""


# ==========================================================
#  REUSABLE UI HELPERS
# ==========================================================
def make_card(title=None):
    card = QFrame()
    card.setObjectName("Card")
    lay = QVBoxLayout(card)
    lay.setContentsMargins(18, 16, 18, 16)
    lay.setSpacing(10)
    if title:
        t = QLabel(title)
        t.setFont(QFont("Segoe UI", 13, QFont.Bold))
        lay.addWidget(t)
    return card, lay


def kpi_card(title, value, color):
    card = QFrame()
    card.setObjectName("KpiCard")
    card.setStyleSheet(
        "background:%s; border-radius:14px;" % color)
    lay = QVBoxLayout(card)
    lay.setContentsMargins(18, 16, 18, 16)
    v = QLabel(str(value))
    v.setStyleSheet("color:white;")
    v.setFont(QFont("Segoe UI", 26, QFont.Bold))
    t = QLabel(title)
    t.setStyleSheet("color:rgba(255,255,255,0.9);")
    t.setFont(QFont("Segoe UI", 11))
    lay.addWidget(v)
    lay.addWidget(t)
    return card, v


# ==========================================================
#  MAIN WINDOW
# ==========================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("%s  v%s" % (APP_NAME, APP_VERSION))
        self.resize(1280, 820)

        self.wb = Workbook()
        self.brain = AIBrain()
        self.dark = False
        self.current_sheet = None

        # TTS engine
        self.tts = None
        if HAVE_TTS:
            try:
                self.tts = pyttsx3.init()
                self.tts.setProperty("rate", 175)
            except Exception:
                self.tts = None

        self._build_ui()
        self.apply_theme()
        self.nav_buttons[0].setChecked(True)
        self.stack.setCurrentIndex(0)

    # ----- speech -----
    def speak(self, text):
        if self.tts:
            try:
                self.tts.say(text)
                self.tts.runAndWait()
            except Exception:
                pass

    # ===== Build layout =====
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ---- Sidebar ----
        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(240)
        sb = QVBoxLayout(sidebar)
        sb.setContentsMargins(10, 6, 10, 10)
        sb.setSpacing(4)

        title = QLabel("📊  Excel AI\nAssistant Pro")
        title.setObjectName("SidebarTitle")
        sb.addWidget(title)

        self.pages = [
            ("🏠  Dashboard", self._page_dashboard),
            ("📁  Upload Excel", self._page_upload),
            ("🛠  Formula Fixer", self._page_fixer),
            ("🧹  Data Cleaner", self._page_cleaner),
            ("✨  Formula Generator", self._page_formula_gen),
            ("📐  Pivot Assistant", self._page_pivot),
            ("⚙️  VBA Generator", self._page_vba),
            ("📈  Dashboard Builder", self._page_dashboard_builder),
            ("💬  AI Chat", self._page_chat),
            ("❤️  Health Check", self._page_health),
            ("📄  Reports", self._page_reports),
            ("🎙  Voice Assistant", self._page_voice),
            ("⚙️  Settings", self._page_settings),
        ]

        self.stack = QStackedWidget()
        self.nav_buttons = []
        for i, (label, builder) in enumerate(self.pages):
            btn = QPushButton(label)
            btn.setObjectName("NavBtn")
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda _, idx=i: self._navigate(idx))
            sb.addWidget(btn)
            self.nav_buttons.append(btn)
            self.stack.addWidget(builder())

        sb.addStretch()
        ver = QLabel("v%s  •  Offline AI" % APP_VERSION)
        ver.setStyleSheet("color:#7f9bc4; padding:8px; font-size:11px;")
        sb.addWidget(ver)

        # ---- Right side ----
        right = QVBoxLayout()
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(0)

        header = QFrame()
        header.setObjectName("Header")
        header.setFixedHeight(64)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(22, 0, 22, 0)
        self.header_title = QLabel("Dashboard")
        self.header_title.setObjectName("HeaderTitle")
        hl.addWidget(self.header_title)
        hl.addStretch()
        self.file_label = QLabel("No file loaded")
        self.file_label.setStyleSheet("color:#7f8fb0;")
        hl.addWidget(self.file_label)
        theme_btn = QPushButton("🌓 Theme")
        theme_btn.setObjectName("Ghost")
        theme_btn.setCursor(Qt.PointingHandCursor)
        theme_btn.clicked.connect(self.toggle_theme)
        hl.addWidget(theme_btn)

        right.addWidget(header)
        right.addWidget(self.stack)

        root.addWidget(sidebar)
        rw = QWidget()
        rw.setLayout(right)
        root.addWidget(rw)

    def _navigate(self, idx):
        for i, b in enumerate(self.nav_buttons):
            b.setChecked(i == idx)
        self.stack.setCurrentIndex(idx)
        label = self.pages[idx][0]
        self.header_title.setText(label.split("  ", 1)[-1])
        # refresh dynamic pages when navigating to them
        if label.endswith("Dashboard"):
            self._refresh_dashboard()
        elif label.endswith("Dashboard Builder"):
            self._refresh_dash_fields()

    # ===== THEME =====
    def apply_theme(self):
        self.setStyleSheet(DARK_QSS if self.dark else LIGHT_QSS)

    def toggle_theme(self):
        self.dark = not self.dark
        self.apply_theme()

    # ----------------------------------------------------
    #  PAGE 1 : DASHBOARD
    # ----------------------------------------------------
    def _page_dashboard(self):
        page = QScrollArea()
        page.setWidgetResizable(True)
        inner = QWidget()
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(18)

        welcome = QLabel("Welcome to Excel AI Assistant Pro")
        welcome.setFont(QFont("Segoe UI", 22, QFont.Bold))
        lay.addWidget(welcome)
        sub = QLabel("Find, explain and fix Excel problems automatically — "
                     "and build formulas, dashboards, VBA & reports in seconds.")
        sub.setStyleSheet("color:#7f8fb0;")
        lay.addWidget(sub)

        # KPI row
        krow = QHBoxLayout()
        self.kpi_sheets = kpi_card("Worksheets", "0", "#2f80ed")
        self.kpi_rows = kpi_card("Total Rows", "0", "#27ae60")
        self.kpi_errors = kpi_card("Errors Found", "0", "#eb5757")
        self.kpi_health = kpi_card("Health Score", "—", "#9b51e0")
        for c, _ in [self.kpi_sheets, self.kpi_rows,
                     self.kpi_errors, self.kpi_health]:
            krow.addWidget(c)
        lay.addLayout(krow)

        # feature grid
        grid_card, gl = make_card("Quick Tools")
        grid = QGridLayout()
        tools = [
            ("📁 Upload Excel", 1), ("🛠 Fix Formula Errors", 2),
            ("🧹 Clean Data", 3), ("✨ Generate Formula", 4),
            ("📐 Pivot Assistant", 5), ("⚙️ VBA Generator", 6),
            ("📈 Build Dashboard", 7), ("💬 Ask AI", 8),
            ("❤️ Health Check", 9), ("📄 Reports", 10),
        ]
        for i, (name, idx) in enumerate(tools):
            b = QPushButton(name)
            b.setObjectName("Ghost")
            b.setMinimumHeight(54)
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, x=idx: self._navigate(x))
            grid.addWidget(b, i // 2, i % 2)
        gl.addLayout(grid)
        lay.addWidget(grid_card)

        lay.addStretch()
        page.setWidget(inner)
        return page

    def _refresh_dashboard(self):
        sheets = len(self.wb.sheets)
        rows = sum(len(df) for df in self.wb.sheets.values())
        errs = len(self.wb.errors)
        self.kpi_sheets[1].setText(str(sheets))
        self.kpi_rows[1].setText(str(rows))
        self.kpi_errors[1].setText(str(errs))
        hc = self.wb.health_check()
        self.kpi_health[1].setText(str(hc["overall"]) if hc else "—")

    # ----------------------------------------------------
    #  PAGE 2 : UPLOAD
    # ----------------------------------------------------
    def _page_upload(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Upload an Excel / CSV file")
        info = QLabel("Supported formats: .xlsx, .xls, .csv — "
                      "the structure is detected automatically.")
        info.setStyleSheet("color:#7f8fb0;")
        cl.addWidget(info)

        btn_row = QHBoxLayout()
        upload_btn = QPushButton("📁  Choose File")
        upload_btn.setObjectName("Primary")
        upload_btn.setCursor(Qt.PointingHandCursor)
        upload_btn.clicked.connect(self.open_file)
        btn_row.addWidget(upload_btn)

        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self._show_sheet)
        btn_row.addWidget(QLabel("Worksheet:"))
        btn_row.addWidget(self.sheet_combo)
        btn_row.addStretch()
        cl.addLayout(btn_row)
        lay.addWidget(card)

        prev_card, pl = make_card("Preview")
        self.preview_table = QTableWidget()
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        pl.addWidget(self.preview_table)
        lay.addWidget(prev_card, 1)
        return page

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel / CSV", "",
            "Spreadsheets (*.xlsx *.xls *.csv);;All Files (*)")
        if not path:
            return
        try:
            self.wb.load(path)
        except Exception as exc:
            QMessageBox.critical(self, "Load error",
                                 "Could not read the file:\n%s" % exc)
            return
        self.file_label.setText("📄 " + os.path.basename(path))
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(list(self.wb.sheets.keys()))
        self.sheet_combo.blockSignals(False)
        if self.wb.sheets:
            self._show_sheet(list(self.wb.sheets.keys())[0])
        self._refresh_dashboard()
        self._refresh_cleaner_combo()
        self._refresh_pivot_combo()
        self._refresh_dash_fields()
        QMessageBox.information(
            self, "Loaded",
            "Loaded %d worksheet(s).\nFound %d potential error(s)."
            % (len(self.wb.sheets), len(self.wb.errors)))

    def _show_sheet(self, name):
        if name not in self.wb.sheets:
            return
        self.current_sheet = name
        df = self.wb.sheets[name]
        self._fill_table(self.preview_table, df.head(200))

    def _fill_table(self, table, df):
        table.clear()
        table.setRowCount(min(len(df), 500))
        table.setColumnCount(len(df.columns))
        table.setHorizontalHeaderLabels([str(c) for c in df.columns])
        for r in range(min(len(df), 500)):
            for c in range(len(df.columns)):
                val = df.iat[r, c]
                item = QTableWidgetItem("" if pd.isna(val) else str(val))
                if str(val).strip() in EXCEL_ERRORS:
                    item.setBackground(QColor("#ffc7ce"))
                    item.setForeground(QColor("#9c0006"))
                table.setItem(r, c, item)
        table.resizeColumnsToContents()

    # ----------------------------------------------------
    #  PAGE 3 : FORMULA FIXER
    # ----------------------------------------------------
    def _page_fixer(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Formula Error Detection")
        desc = QLabel("Detects #VALUE!, #DIV/0!, #N/A, #REF!, #NAME?, #NUM!, "
                      "#NULL!, circular & broken references.\n"
                      "STEP 1: Scan to review errors yourself.   "
                      "STEP 2: Auto-Fix & Save to get a corrected file.")
        desc.setStyleSheet("color:#7f8fb0;")
        cl.addWidget(desc)
        row = QHBoxLayout()
        scan_btn = QPushButton("🔍  Scan Workbook")
        scan_btn.setObjectName("Primary")
        scan_btn.setCursor(Qt.PointingHandCursor)
        scan_btn.clicked.connect(self._run_scan)
        row.addWidget(scan_btn)
        autofix_btn = QPushButton("⚡  Auto-Fix & Save File")
        autofix_btn.setObjectName("Success")
        autofix_btn.setCursor(Qt.PointingHandCursor)
        autofix_btn.clicked.connect(self._auto_fix_save)
        row.addWidget(autofix_btn)
        fix_btn = QPushButton("💡  How to Fix (Tips)")
        fix_btn.setObjectName("Ghost")
        fix_btn.setCursor(Qt.PointingHandCursor)
        fix_btn.clicked.connect(self._auto_fix)
        row.addWidget(fix_btn)
        row.addStretch()
        cl.addLayout(row)
        lay.addWidget(card)

        res_card, rl = make_card("Detected Errors")
        self.error_table = QTableWidget()
        self.error_table.setColumnCount(5)
        self.error_table.setHorizontalHeaderLabels(
            ["Sheet", "Location", "Error", "Reason", "Suggested Fix"])
        self.error_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.Stretch)
        rl.addWidget(self.error_table)
        lay.addWidget(res_card, 1)
        return page

    def _run_scan(self):
        if not self.wb.sheets:
            QMessageBox.warning(self, "No file", "Please upload a file first.")
            return
        self.wb._scan_errors()
        errs = self.wb.errors
        self.error_table.setRowCount(len(errs))
        for i, e in enumerate(errs):
            for j, key in enumerate(["sheet", "cell", "error",
                                     "reason", "fix"]):
                item = QTableWidgetItem(e[key])
                if key == "error":
                    item.setForeground(QColor("#eb5757"))
                    item.setFont(QFont("Segoe UI", 10, QFont.Bold))
                self.error_table.setItem(i, j, item)
        self.error_table.resizeColumnsToContents()
        self.error_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.Stretch)
        self._refresh_dashboard()
        if not errs:
            QMessageBox.information(self, "Clean!",
                                    "No errors found. Your workbook is clean. 🎉")

    def _auto_fix(self):
        if not self.wb.errors:
            QMessageBox.information(self, "Nothing to fix",
                                    "Run a scan first, or no errors exist.")
            return
        msg = "Suggested one-click fixes:\n\n"
        seen = set()
        for e in self.wb.errors:
            if e["error"] in seen:
                continue
            seen.add(e["error"])
            msg += "• %s\n   %s\n\n" % (e["error"], e["fix"])
        QMessageBox.information(self, "Auto-Fix Suggestions", msg)

    def _auto_fix_save(self):
        """Automatically repair errors and save a corrected copy of the file."""
        if not self.wb.sheets:
            QMessageBox.warning(self, "No file", "Please upload a file first.")
            return
        # default name: <original>_FIXED.xlsx
        base = "fixed_workbook"
        if self.wb.path:
            base = os.path.splitext(os.path.basename(self.wb.path))[0] + "_FIXED"
        default = base + (".csv" if (self.wb.path or "").lower().endswith(".csv")
                          else ".xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save fixed file", default,
            "Excel (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        try:
            fixed, manual = self.wb.auto_fix(path)
        except Exception as exc:
            QMessageBox.critical(self, "Auto-Fix error",
                                 "Could not fix the file:\n%s" % exc)
            return

        msg = "✅ Fixed %d error(s) and saved a corrected copy:\n%s\n\n" \
              % (len(fixed), path)
        if fixed:
            msg += "What was fixed:\n"
            for f in fixed[:15]:
                msg += "  • %s!%s  (%s) -> %s\n" % (
                    f["sheet"], f["cell"], f["error"], f["action"])
            if len(fixed) > 15:
                msg += "  ... and %d more\n" % (len(fixed) - 15)
        if manual:
            msg += ("\n⚠️ %d item(s) need manual review (e.g. circular "
                    "references):\n" % len(manual))
            for f in manual[:8]:
                msg += "  • %s!%s  (%s)\n" % (f["sheet"], f["cell"], f["error"])
        if not fixed and not manual:
            msg = "No fixable errors were found. Your file is already clean. 🎉"
        QMessageBox.information(self, "Auto-Fix Complete", msg)

    # ----------------------------------------------------
    #  PAGE 4 : DATA CLEANER
    # ----------------------------------------------------
    def _page_cleaner(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Data Cleaning Options")
        top = QHBoxLayout()
        top.addWidget(QLabel("Worksheet:"))
        self.clean_combo = QComboBox()
        top.addWidget(self.clean_combo)
        top.addStretch()
        cl.addLayout(top)

        self.clean_checks = {}
        opts = [
            ("duplicates", "Remove duplicate rows"),
            ("blanks", "Remove blank rows"),
            ("trim", "Trim extra spaces"),
            ("title", "Standardise text (Title Case)"),
            ("dates", "Fix date formats"),
            ("columns", "Standardise column headers"),
        ]
        grid = QGridLayout()
        for i, (key, label) in enumerate(opts):
            cb = QCheckBox(label)
            cb.setChecked(True)
            self.clean_checks[key] = cb
            grid.addWidget(cb, i // 2, i % 2)
        cl.addLayout(grid)

        run = QPushButton("🧹  Clean Now")
        run.setObjectName("Success")
        run.setCursor(Qt.PointingHandCursor)
        run.clicked.connect(self._run_clean)
        cl.addWidget(run)
        lay.addWidget(card)

        rep_card, rl = make_card("Cleaning Report")
        self.clean_report = QPlainTextEdit()
        self.clean_report.setReadOnly(True)
        rl.addWidget(self.clean_report)
        save_btn = QPushButton("💾  Save Cleaned File")
        save_btn.setObjectName("Primary")
        save_btn.setCursor(Qt.PointingHandCursor)
        save_btn.clicked.connect(self._save_cleaned)
        rl.addWidget(save_btn)
        lay.addWidget(rep_card, 1)

        self._cleaned_df = None
        return page

    def _refresh_cleaner_combo(self):
        self.clean_combo.clear()
        self.clean_combo.addItems(list(self.wb.sheets.keys()))

    def _run_clean(self):
        name = self.clean_combo.currentText()
        if name not in self.wb.sheets:
            QMessageBox.warning(self, "No file", "Please upload a file first.")
            return
        opts = {k: cb.isChecked() for k, cb in self.clean_checks.items()}
        df = self.wb.sheets[name].copy()
        cleaned, report = self.wb.clean(df, opts)
        self._cleaned_df = cleaned
        self.clean_report.setPlainText("CLEANING REPORT\n" + "=" * 40 + "\n"
                                       + "\n".join("• " + r for r in report))

    def _save_cleaned(self):
        if self._cleaned_df is None:
            QMessageBox.warning(self, "Nothing", "Run 'Clean Now' first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save cleaned file", "cleaned_data.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)")
        if not path:
            return
        try:
            if path.lower().endswith(".csv"):
                self._cleaned_df.to_csv(path, index=False)
            else:
                self._cleaned_df.to_excel(path, index=False)
            QMessageBox.information(self, "Saved", "Cleaned file saved:\n%s"
                                    % path)
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))

    # ----------------------------------------------------
    #  PAGE 5 : FORMULA GENERATOR
    # ----------------------------------------------------
    def _page_formula_gen(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("AI Formula Generator")
        cl.addWidget(QLabel("Describe what you need in plain English:"))
        row = QHBoxLayout()
        self.formula_input = QLineEdit()
        self.formula_input.setPlaceholderText(
            "e.g. calculate tax, find top 10 sales, lookup customer data...")
        self.formula_input.returnPressed.connect(self._gen_formula)
        row.addWidget(self.formula_input)
        gen = QPushButton("✨  Generate")
        gen.setObjectName("Primary")
        gen.setCursor(Qt.PointingHandCursor)
        gen.clicked.connect(self._gen_formula)
        row.addWidget(gen)
        cl.addLayout(row)

        chips = QHBoxLayout()
        for ex in ["Calculate tax", "Find top 10 sales",
                   "Lookup customer data", "Create attendance tracker"]:
            b = QPushButton(ex)
            b.setObjectName("Ghost")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, t=ex: (
                self.formula_input.setText(t), self._gen_formula()))
            chips.addWidget(b)
        cl.addLayout(chips)
        lay.addWidget(card)

        out_card, ol = make_card("Result")
        self.formula_out = QPlainTextEdit()
        self.formula_out.setReadOnly(True)
        ol.addWidget(self.formula_out)
        copy = QPushButton("📋  Copy Formula")
        copy.setObjectName("Success")
        copy.setCursor(Qt.PointingHandCursor)
        copy.clicked.connect(lambda: self._copy(self._last_formula))
        ol.addWidget(copy)
        lay.addWidget(out_card, 1)
        self._last_formula = ""
        return page

    def _gen_formula(self):
        req = self.formula_input.text().strip()
        if not req:
            return
        res = self.brain.generate_formula(req)
        self._last_formula = res["formula"]
        self.formula_out.setPlainText(
            "FORMULA\n%s\n%s\n\nEXPLANATION\n%s\n\nSAMPLE OUTPUT\n%s"
            % ("-" * 40, res["formula"], res["explanation"], res["sample"]))

    # ----------------------------------------------------
    #  PAGE 6 : PIVOT ASSISTANT
    # ----------------------------------------------------
    def _page_pivot(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Pivot Table Assistant")
        grid = QGridLayout()
        grid.addWidget(QLabel("Worksheet:"), 0, 0)
        self.pivot_sheet = QComboBox()
        self.pivot_sheet.currentTextChanged.connect(self._refresh_pivot_fields)
        grid.addWidget(self.pivot_sheet, 0, 1)
        grid.addWidget(QLabel("Group by (Rows):"), 1, 0)
        self.pivot_row = QComboBox()
        grid.addWidget(self.pivot_row, 1, 1)
        grid.addWidget(QLabel("Value:"), 2, 0)
        self.pivot_val = QComboBox()
        grid.addWidget(self.pivot_val, 2, 1)
        grid.addWidget(QLabel("Function:"), 3, 0)
        self.pivot_func = QComboBox()
        self.pivot_func.addItems(["sum", "mean", "count", "max", "min"])
        grid.addWidget(self.pivot_func, 3, 1)
        cl.addLayout(grid)
        build = QPushButton("📐  Build Pivot Table")
        build.setObjectName("Primary")
        build.setCursor(Qt.PointingHandCursor)
        build.clicked.connect(self._build_pivot)
        cl.addWidget(build)
        lay.addWidget(card)

        res_card, rl = make_card("Pivot Result")
        self.pivot_table = QTableWidget()
        rl.addWidget(self.pivot_table)
        lay.addWidget(res_card, 1)
        return page

    def _refresh_pivot_combo(self):
        self.pivot_sheet.clear()
        self.pivot_sheet.addItems(list(self.wb.sheets.keys()))

    def _refresh_pivot_fields(self, name):
        if name not in self.wb.sheets:
            return
        df = self.wb.sheets[name]
        cols = [str(c) for c in df.columns]
        nums = [str(c) for c in df.select_dtypes(include=np.number).columns]
        self.pivot_row.clear(); self.pivot_row.addItems(cols)
        self.pivot_val.clear(); self.pivot_val.addItems(nums or cols)

    def _build_pivot(self):
        name = self.pivot_sheet.currentText()
        if name not in self.wb.sheets:
            QMessageBox.warning(self, "No file", "Upload a file first.")
            return
        df = self.wb.sheets[name]
        rkey = self.pivot_row.currentText()
        vkey = self.pivot_val.currentText()
        func = self.pivot_func.currentText()
        try:
            pivot = pd.pivot_table(df, index=rkey, values=vkey,
                                   aggfunc=func).reset_index()
        except Exception as exc:
            QMessageBox.critical(self, "Pivot error", str(exc))
            return
        self._fill_table(self.pivot_table, pivot)

    # ----------------------------------------------------
    #  PAGE 7 : VBA GENERATOR
    # ----------------------------------------------------
    def _page_vba(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("VBA & Macro Generator")
        cl.addWidget(QLabel("Describe the macro you want in plain English:"))
        row = QHBoxLayout()
        self.vba_input = QLineEdit()
        self.vba_input.setPlaceholderText(
            "e.g. create employee attendance report, remove duplicates, "
            "highlight error cells, email workbook...")
        row.addWidget(self.vba_input)
        gen = QPushButton("⚙️  Generate VBA")
        gen.setObjectName("Primary")
        gen.setCursor(Qt.PointingHandCursor)
        gen.clicked.connect(self._gen_vba)
        row.addWidget(gen)
        cl.addLayout(row)
        lay.addWidget(card)

        out_card, ol = make_card("Generated Macro")
        self.vba_out = QPlainTextEdit()
        self.vba_out.setReadOnly(True)
        self.vba_out.setFont(QFont("Consolas", 10))
        ol.addWidget(self.vba_out)
        copy = QPushButton("📋  Copy Code")
        copy.setObjectName("Success")
        copy.setCursor(Qt.PointingHandCursor)
        copy.clicked.connect(lambda: self._copy(self.vba_out.toPlainText()))
        ol.addWidget(copy)
        lay.addWidget(out_card, 1)
        return page

    def _gen_vba(self):
        req = self.vba_input.text().strip() or "custom macro"
        code = self.brain.generate_vba(req)
        explain = ("\n\n' ---- HOW TO USE ----\n"
                   "' 1. Press ALT+F11 in Excel to open the VBA editor.\n"
                   "' 2. Insert > Module, then paste this code.\n"
                   "' 3. Press F5 to run, or assign it to a button.")
        self.vba_out.setPlainText(code + explain)

    # ----------------------------------------------------
    #  PAGE 8 : DASHBOARD BUILDER
    # ----------------------------------------------------
    def _page_dashboard_builder(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Dashboard Generator")
        row = QHBoxLayout()
        row.addWidget(QLabel("Worksheet:"))
        self.dash_sheet = QComboBox()
        row.addWidget(self.dash_sheet)
        row.addWidget(QLabel("Category:"))
        self.dash_cat = QComboBox()
        row.addWidget(self.dash_cat)
        row.addWidget(QLabel("Value:"))
        self.dash_val = QComboBox()
        row.addWidget(self.dash_val)
        build = QPushButton("📈  Generate Dashboard")
        build.setObjectName("Primary")
        build.setCursor(Qt.PointingHandCursor)
        build.clicked.connect(self._build_dashboard)
        row.addWidget(build)
        self.dash_sheet.currentTextChanged.connect(self._refresh_dash_fields)
        cl.addLayout(row)
        lay.addWidget(card)

        self.dash_canvas = FigureCanvas(Figure(figsize=(8, 5)))
        chart_card, chl = make_card("Dashboard")
        chl.addWidget(self.dash_canvas)
        lay.addWidget(chart_card, 1)
        return page

    def _refresh_dash_fields(self, name=None):
        # called on navigation too
        self.dash_sheet.blockSignals(True)
        current = self.dash_sheet.currentText()
        if self.dash_sheet.count() != len(self.wb.sheets):
            self.dash_sheet.clear()
            self.dash_sheet.addItems(list(self.wb.sheets.keys()))
        self.dash_sheet.blockSignals(False)
        name = name or self.dash_sheet.currentText()
        if name in self.wb.sheets:
            df = self.wb.sheets[name]
            cols = [str(c) for c in df.columns]
            nums = [str(c) for c in df.select_dtypes(include=np.number).columns]
            self.dash_cat.clear(); self.dash_cat.addItems(cols)
            self.dash_val.clear(); self.dash_val.addItems(nums or cols)

    def _build_dashboard(self):
        name = self.dash_sheet.currentText()
        if name not in self.wb.sheets:
            QMessageBox.warning(self, "No file", "Upload a file first.")
            return
        df = self.wb.sheets[name]
        cat = self.dash_cat.currentText()
        val = self.dash_val.currentText()
        try:
            grouped = df.groupby(cat)[val].sum().sort_values(
                ascending=False).head(10)
        except Exception as exc:
            QMessageBox.critical(self, "Error", str(exc))
            return

        fig = self.dash_canvas.figure
        fig.clear()
        colors = ["#2f80ed", "#27ae60", "#9b51e0", "#f2994a", "#eb5757"]

        ax1 = fig.add_subplot(221)
        grouped.plot(kind="bar", ax=ax1, color="#2f80ed")
        ax1.set_title("Top by %s" % val, fontsize=10)
        ax1.tick_params(axis="x", labelsize=7, rotation=45)

        ax2 = fig.add_subplot(222)
        grouped.head(5).plot(kind="pie", ax=ax2, autopct="%1.0f%%",
                             colors=colors, textprops={"fontsize": 7})
        ax2.set_ylabel("")
        ax2.set_title("Share (Top 5)", fontsize=10)

        ax3 = fig.add_subplot(223)
        grouped.plot(kind="line", ax=ax3, marker="o", color="#27ae60")
        ax3.set_title("Trend", fontsize=10)
        ax3.tick_params(axis="x", labelsize=7, rotation=45)

        ax4 = fig.add_subplot(224)
        ax4.axis("off")
        total = grouped.sum()
        ax4.text(0.5, 0.7, "{:,.0f}".format(total), ha="center",
                 fontsize=22, color="#1b3a6b", fontweight="bold")
        ax4.text(0.5, 0.4, "Total %s" % val, ha="center", fontsize=10)
        ax4.text(0.5, 0.2, "Top: %s" % str(grouped.index[0]),
                 ha="center", fontsize=9, color="#7f8fb0")

        fig.tight_layout()
        self.dash_canvas.draw()

    # ----------------------------------------------------
    #  PAGE 9 : AI CHAT
    # ----------------------------------------------------
    def _page_chat(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(12)

        self.chat_view = QTextEdit()
        self.chat_view.setReadOnly(True)
        self._chat_append("AI", "Hi! I'm your Excel AI Assistant. Ask me "
                          "anything — formulas, errors, VLOOKUP, XLOOKUP, "
                          "pivot tables, Power Query, VBA and more.")
        lay.addWidget(self.chat_view, 1)

        row = QHBoxLayout()
        self.chat_input = QLineEdit()
        self.chat_input.setPlaceholderText("Type your question...")
        self.chat_input.returnPressed.connect(self._send_chat)
        row.addWidget(self.chat_input)
        send = QPushButton("Send")
        send.setObjectName("Primary")
        send.setCursor(Qt.PointingHandCursor)
        send.clicked.connect(self._send_chat)
        row.addWidget(send)
        lay.addLayout(row)
        return page

    def _chat_append(self, who, text):
        color = "#2f80ed" if who == "You" else "#27ae60"
        safe = text.replace("\n", "<br>")
        self.chat_view.append(
            "<p><b style='color:%s'>%s:</b> %s</p>" % (color, who, safe))

    def _send_chat(self):
        msg = self.chat_input.text().strip()
        if not msg:
            return
        self._chat_append("You", msg)
        self.chat_input.clear()
        reply = self.brain.chat(msg)
        self._chat_append("AI", reply)

    # ----------------------------------------------------
    #  PAGE 10 : HEALTH CHECK
    # ----------------------------------------------------
    def _page_health(self):
        page = QScrollArea()
        page.setWidgetResizable(True)
        inner = QWidget()
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Excel Health Check")
        run = QPushButton("❤️  Run Health Check")
        run.setObjectName("Primary")
        run.setCursor(Qt.PointingHandCursor)
        run.clicked.connect(self._run_health)
        cl.addWidget(run)
        lay.addWidget(card)

        self.health_bars = {}
        bars_card, bl = make_card("Scores")
        for key, label in [("overall", "Overall Score"),
                           ("error", "Error Score"),
                           ("quality", "Data Quality"),
                           ("complexity", "Formula Simplicity"),
                           ("performance", "Performance")]:
            bl.addWidget(QLabel(label))
            bar = QProgressBar()
            bar.setMaximum(100)
            self.health_bars[key] = bar
            bl.addWidget(bar)
        lay.addWidget(bars_card)

        tips_card, tl = make_card("Improvement Suggestions")
        self.health_tips = QPlainTextEdit()
        self.health_tips.setReadOnly(True)
        tl.addWidget(self.health_tips)
        lay.addWidget(tips_card)

        lay.addStretch()
        page.setWidget(inner)
        return page

    def _run_health(self):
        hc = self.wb.health_check()
        if not hc:
            QMessageBox.warning(self, "No file", "Upload a file first.")
            return
        for key, bar in self.health_bars.items():
            bar.setValue(hc[key])
        self.health_tips.setPlainText(
            "STATS: %s\n\nSUGGESTIONS:\n%s"
            % (hc["stats"], "\n".join("• " + t for t in hc["tips"])))

    # ----------------------------------------------------
    #  PAGE 11 : REPORTS
    # ----------------------------------------------------
    def _page_reports(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Report Generator")
        cl.addWidget(QLabel("Generate a full audit & data-quality report "
                            "from the loaded workbook."))
        row = QHBoxLayout()
        gen = QPushButton("📄  Generate Report")
        gen.setObjectName("Primary")
        gen.setCursor(Qt.PointingHandCursor)
        gen.clicked.connect(self._gen_report)
        row.addWidget(gen)
        save_txt = QPushButton("💾  Save as Text")
        save_txt.setObjectName("Ghost")
        save_txt.setCursor(Qt.PointingHandCursor)
        save_txt.clicked.connect(lambda: self._save_report("txt"))
        row.addWidget(save_txt)
        save_xl = QPushButton("💾  Save as Excel")
        save_xl.setObjectName("Ghost")
        save_xl.setCursor(Qt.PointingHandCursor)
        save_xl.clicked.connect(lambda: self._save_report("xlsx"))
        row.addWidget(save_xl)
        row.addStretch()
        cl.addLayout(row)
        lay.addWidget(card)

        out_card, ol = make_card("Report Preview")
        self.report_view = QPlainTextEdit()
        self.report_view.setReadOnly(True)
        ol.addWidget(self.report_view)
        lay.addWidget(out_card, 1)
        return page

    def _build_report_text(self):
        hc = self.wb.health_check()
        lines = []
        lines.append("=" * 56)
        lines.append("        EXCEL AI ASSISTANT PRO - AUDIT REPORT")
        lines.append("=" * 56)
        lines.append("Generated: %s" % datetime.datetime.now()
                     .strftime("%Y-%m-%d %H:%M"))
        lines.append("File: %s" % (os.path.basename(self.wb.path)
                                    if self.wb.path else "—"))
        lines.append("")
        lines.append("WORKBOOK STRUCTURE")
        lines.append("-" * 56)
        for name, df in self.wb.sheets.items():
            lines.append("  Sheet '%s': %d rows x %d cols"
                         % (name, len(df), len(df.columns)))
        lines.append("")
        if hc:
            lines.append("HEALTH SCORES")
            lines.append("-" * 56)
            lines.append("  Overall:      %d/100" % hc["overall"])
            lines.append("  Error:        %d/100" % hc["error"])
            lines.append("  Data Quality: %d/100" % hc["quality"])
            lines.append("  Simplicity:   %d/100" % hc["complexity"])
            lines.append("  Performance:  %d/100" % hc["performance"])
            lines.append("")
            lines.append("  Stats: %s" % hc["stats"])
            lines.append("")
        lines.append("DETECTED ERRORS (%d)" % len(self.wb.errors))
        lines.append("-" * 56)
        if self.wb.errors:
            for e in self.wb.errors:
                lines.append("  [%s] %s @ %s - %s"
                             % (e["error"], e["sheet"], e["cell"], e["fix"]))
        else:
            lines.append("  None - workbook is clean.")
        lines.append("")
        if hc:
            lines.append("RECOMMENDATIONS")
            lines.append("-" * 56)
            for t in hc["tips"]:
                lines.append("  • " + t)
        lines.append("")
        lines.append("=" * 56)
        lines.append("End of report.")
        return "\n".join(lines)

    def _gen_report(self):
        if not self.wb.sheets:
            QMessageBox.warning(self, "No file", "Upload a file first.")
            return
        self.report_view.setPlainText(self._build_report_text())

    def _save_report(self, fmt):
        if not self.wb.sheets:
            QMessageBox.warning(self, "No file", "Upload a file first.")
            return
        text = self._build_report_text()
        if fmt == "txt":
            path, _ = QFileDialog.getSaveFileName(
                self, "Save report", "excel_audit_report.txt",
                "Text (*.txt)")
            if not path:
                return
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
        else:
            path, _ = QFileDialog.getSaveFileName(
                self, "Save report", "excel_audit_report.xlsx",
                "Excel (*.xlsx)")
            if not path:
                return
            err_df = pd.DataFrame(self.wb.errors) if self.wb.errors \
                else pd.DataFrame([{"info": "No errors found"}])
            summary = pd.DataFrame(
                [{"Sheet": n, "Rows": len(d), "Columns": len(d.columns)}
                 for n, d in self.wb.sheets.items()])
            with pd.ExcelWriter(path) as writer:
                summary.to_excel(writer, sheet_name="Summary", index=False)
                err_df.to_excel(writer, sheet_name="Errors", index=False)
        QMessageBox.information(self, "Saved", "Report saved:\n%s" % path)

    # ----------------------------------------------------
    #  PAGE 12 : VOICE ASSISTANT
    # ----------------------------------------------------
    def _page_voice(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Voice Assistant")
        status = "🔊 Voice output: ENABLED" if self.tts \
            else "🔇 Voice output not available (pyttsx3 missing)"
        cl.addWidget(QLabel(status))
        cl.addWidget(QLabel("Type a command and the assistant will respond "
                            "with text + voice:"))
        row = QHBoxLayout()
        self.voice_input = QLineEdit()
        self.voice_input.setPlaceholderText(
            "e.g. find errors in my workbook, create a tax formula...")
        self.voice_input.returnPressed.connect(self._voice_command)
        row.addWidget(self.voice_input)
        speak = QPushButton("🎙  Run & Speak")
        speak.setObjectName("Primary")
        speak.setCursor(Qt.PointingHandCursor)
        speak.clicked.connect(self._voice_command)
        row.addWidget(speak)
        cl.addLayout(row)

        chips = QHBoxLayout()
        for ex in ["Find errors in my workbook",
                   "Create a tax formula",
                   "Generate attendance dashboard"]:
            b = QPushButton(ex)
            b.setObjectName("Ghost")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, t=ex: (
                self.voice_input.setText(t), self._voice_command()))
            chips.addWidget(b)
        cl.addLayout(chips)
        lay.addWidget(card)

        out_card, ol = make_card("Assistant Response")
        self.voice_out = QPlainTextEdit()
        self.voice_out.setReadOnly(True)
        ol.addWidget(self.voice_out)
        lay.addWidget(out_card, 1)
        return page

    def _voice_command(self):
        cmd = self.voice_input.text().strip()
        if not cmd:
            return
        low = cmd.lower()
        if "error" in low:
            self.wb._scan_errors()
            n = len(self.wb.errors)
            resp = ("I scanned your workbook and found %d error(s)." % n
                    if self.wb.sheets else
                    "Please upload a workbook first, then I can scan it.")
        elif "tax" in low or "formula" in low:
            res = self.brain.generate_formula(cmd)
            resp = "Here is a formula: %s. %s" % (res["formula"].split(chr(10))[0],
                                                  res["explanation"])
        elif "dashboard" in low:
            resp = ("Open the Dashboard Builder tab and pick a category and "
                    "value — I'll generate KPI, bar, pie and trend charts.")
        else:
            resp = self.brain.chat(cmd)
        self.voice_out.setPlainText(resp)
        self.speak(resp)

    # ----------------------------------------------------
    #  PAGE 13 : SETTINGS
    # ----------------------------------------------------
    def _page_settings(self):
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        card, cl = make_card("Settings & User Profile")
        cl.addWidget(QLabel("Display name:"))
        self.set_name = QLineEdit("Excel Power User")
        cl.addWidget(self.set_name)

        cl.addWidget(QLabel("Theme:"))
        self.set_theme = QComboBox()
        self.set_theme.addItems(["Light", "Dark"])
        self.set_theme.currentTextChanged.connect(
            lambda t: self._set_theme(t == "Dark"))
        cl.addWidget(self.set_theme)

        cl.addWidget(QLabel("OpenAI API Key (optional — enables cloud GPT "
                            "answers; leave blank to use built-in offline AI):"))
        self.set_api = QLineEdit()
        self.set_api.setEchoMode(QLineEdit.Password)
        self.set_api.setPlaceholderText("sk-...")
        cl.addWidget(self.set_api)

        self.voice_toggle = QCheckBox("Enable voice responses")
        self.voice_toggle.setChecked(bool(self.tts))
        cl.addWidget(self.voice_toggle)

        save = QPushButton("💾  Save Settings")
        save.setObjectName("Success")
        save.setCursor(Qt.PointingHandCursor)
        save.clicked.connect(lambda: QMessageBox.information(
            self, "Saved", "Settings saved for this session."))
        cl.addWidget(save)
        lay.addWidget(card)

        about_card, al = make_card("About")
        al.addWidget(QLabel(
            "%s v%s\n\nA complete offline Excel troubleshooting & "
            "productivity assistant.\nBuilt with Python + PyQt5.\n\n"
            "Features: error detection & fixes, formula & VBA generation, "
            "data cleaning, pivot & dashboard builders, AI chat, health "
            "check, reports and voice." % (APP_NAME, APP_VERSION)))
        lay.addWidget(about_card)
        lay.addStretch()
        return page

    def _set_theme(self, dark):
        self.dark = dark
        self.apply_theme()

    # ----- shared helpers -----
    def _copy(self, text):
        if not text:
            return
        QApplication.clipboard().setText(text)
        QMessageBox.information(self, "Copied", "Copied to clipboard!")


# ==========================================================
#  ENTRY POINT
# ==========================================================
def main():
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    try:
        app.setFont(QFont("Segoe UI", 10))
    except Exception:
        pass
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        input("\nAn error occurred. Press Enter to exit...")
